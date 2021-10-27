import os
import time as t
import openpyxl as op
from openpyxl import workbook
import funciones as f
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

ComboS = 650
ComboD = 700
ComboT = 800
Flurby = 250


os.system("cls")


wb = op.Workbook()

try:
	wb = load_workbook("Pedidos.xlsx")
except:
	wb = Workbook()
	ws1 = workbook.active

	Cel1 = ws1["A1"] = "Cliente"
	Cel1 = Alignment(horizontal='center')

	Cel2 = ws1["B1"] = "Fecha"
	Cel2 = Alignment(horizontal='center')
	Cel2 = ws1.column_dimensions['B'].width = 25

	Cel3 = ws1["C1"] = "Combo S"
	Cel3 = Alignment(horizontal='center')

	Cel4 = ws1["D1"] = "Combo D"
	Cel4 = Alignment(horizontal='center')

	Cel5 = ws1["E1"] = "Combo T"
	Cel5 = Alignment(horizontal='center')

	Cel6 = ws1["F1"] = "Flurby"
	Cel6 = Alignment(horizontal='center')

	Cel7 = ws1["G1"] = "Total"
	Cel7 = Alignment(horizontal='center')

	wb.save(filename="pedidos.xlsx")

ws = wb.active


while True:
	print("Ingrese nombre del cliente: ")
	Cliente = input(">>> ")
	Cliente = f.verificar_vacio(Cliente)
	Cliente = f.verificar_nombre(Cliente.capitalize())

	Cant_comboS = input("Ingrese la cantidad de comboS: \n>>> ")
	Cant_comboS = f.verificar_vacio(Cant_comboS)
	Cant_comboS = f.convertir(Cant_comboS)
	
	Cant_comboD = input("Ingrese la cantidad de comboD: \n>>> ")
	Cant_comboD = f.verificar_vacio(Cant_comboD)
	Cant_comboD = f.convertir(Cant_comboD)
	
	Cant_comboT = input("Ingrese la cantidad de comboT: \n>>> ")
	Cant_comboT = f.verificar_vacio(Cant_comboT)
	Cant_comboT = f.convertir(Cant_comboT)
		
	Cant_Flurby = input("Ingrese la cantidad de Flurby: \n>>> ")
	Cant_Flurby = f.verificar_vacio(Cant_Flurby)
	Cant_Flurby = f.convertir(Cant_Flurby)

	Total = Cant_comboS * ComboS + Cant_comboD * ComboD + Cant_comboT * ComboT + Cant_Flurby * Flurby

	print("El total es $", Total)

	Abono = input("Con cuanto desea abonar: \n>>> ")
	Abono = f.verificar_vacio(Abono)
	Abono = f.verificar_numero_str(Abono)
	Abono = f.convertir(Abono)
	Abono = f.verif_vuelto(Total, Abono)

	vuelto = Abono - Total

	f.total_en_caja += (Abono - vuelto)

	Fecha = t.asctime()

	ws.append([Cliente, Fecha, Cant_comboS, Cant_comboD, Cant_comboT, Cant_Flurby, Total])
	wb.save("Pedidos.xlsx")
	print("Su vuelto es de $", vuelto)

	t.sleep(3)

	break